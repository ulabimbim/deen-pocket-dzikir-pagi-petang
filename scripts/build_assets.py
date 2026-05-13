import json
import re
import shutil
from collections import Counter
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "content" / "dzikir-pagi-petang.xlsx"
DATA_JSON = ROOT / "data" / "dzikir.json"
DOWNLOAD_XLSX = Path("/Users/widirosa/Downloads/Dzikir Pagi Petang - Deen Pocket (3).xlsx")

REQUIRED_HEADERS = [
    "Sesi",
    "Urutan Pagi",
    "Urutan Petang",
    "Judul",
    "Arab",
    "Latin",
    "Arti",
    "Jumlah",
    "Tipe",
    "Sumber",
    "Keterangan",
]


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def slugify(value):
    slug = re.sub(r"[^a-z0-9]+", "-", cell_text(value).lower()).strip("-")
    return slug or "dzikir"


def unique_slug(base, seen):
    if base not in seen:
        seen[base] = 1
        return base

    seen[base] += 1
    return f"{base}-{seen[base]}"


def normalize_sessions(value):
    text = cell_text(value).lower()
    text = text.replace("&", " dan ").replace("/", ",")
    sessions = []
    if "pagi" in text:
        sessions.append("pagi")
    if "petang" in text or "sore" in text:
        sessions.append("petang")
    return sessions


def normalize_order(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        if int(value) == value:
            return int(value)
        return value
    text = cell_text(value).strip()
    if not text:
        return None
    number = float(text)
    if int(number) == number:
        return int(number)
    return number


def normalize_count(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        if int(value) == value:
            return int(value)
        return value
    text = cell_text(value).strip()
    if text.endswith("x"):
        text = text[:-1].strip()
    number = float(text)
    if int(number) == number:
        return int(number)
    return number


def normalize_type(value, count):
    raw = cell_text(value).strip().lower()
    if raw == "next":
        return "read"
    if raw in ("read", "counter"):
        return raw
    if not raw:
        return "counter" if count and count > 1 else "read"
    return raw


def validate_items(items):
    warnings = []

    for item in items:
        if "pagi" in item["sessions"] and item["morningOrder"] is None:
            warnings.append(f"{item['id']}: item pagi tanpa Urutan Pagi")
        if "petang" in item["sessions"] and item["eveningOrder"] is None:
            warnings.append(f"{item['id']}: item petang tanpa Urutan Petang")
        if not isinstance(item["count"], (int, float)):
            warnings.append(f"{item['id']}: Jumlah bukan number")
        if item["type"] not in ("read", "counter"):
            warnings.append(f"{item['id']}: Tipe tidak valid ({item['type']})")

    morning_orders = [
        item["morningOrder"]
        for item in items
        if "pagi" in item["sessions"] and item["morningOrder"] is not None
    ]
    evening_orders = [
        item["eveningOrder"]
        for item in items
        if "petang" in item["sessions"] and item["eveningOrder"] is not None
    ]

    for order, count in Counter(morning_orders).items():
        if count > 1:
            warnings.append(f"Duplikasi Urutan Pagi: {order}")
    for order, count in Counter(evening_orders).items():
        if count > 1:
            warnings.append(f"Duplikasi Urutan Petang: {order}")

    return warnings


def build_dzikir_json():
    if not SOURCE_XLSX.exists():
        if DOWNLOAD_XLSX.exists():
            SOURCE_XLSX.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(DOWNLOAD_XLSX, SOURCE_XLSX)
        else:
            raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")

    workbook = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}

    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_index]
    if missing_headers:
        raise ValueError(f"Missing required XLSX columns: {', '.join(missing_headers)}")

    seen_slugs = {}
    items = []
    source_fields = REQUIRED_HEADERS + ["Catatan", "Note"]

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source = {
            field: row[header_index[field]]
            for field in source_fields
            if field in header_index
        }
        if not any(source.get(field) not in (None, "") for field in REQUIRED_HEADERS):
            continue

        count = normalize_count(source.get("Jumlah"))
        item_type = normalize_type(source.get("Tipe"), count)
        title = cell_text(source.get("Judul"))

        item = {
            "id": unique_slug(slugify(title), seen_slugs),
            "sessions": normalize_sessions(source.get("Sesi")),
            "morningOrder": normalize_order(source.get("Urutan Pagi")),
            "eveningOrder": normalize_order(source.get("Urutan Petang")),
            "title": title,
            "arabic": cell_text(source.get("Arab")),
            "latin": cell_text(source.get("Latin")),
            "translation": cell_text(source.get("Arti")),
            "count": count,
            "type": item_type,
            "source": cell_text(source.get("Sumber")),
            "description": cell_text(source.get("Keterangan")),
            "note": cell_text(source.get("Catatan") or source.get("Note")),
            "sourceRow": row_number,
        }
        items.append(item)

    warnings = validate_items(items)
    DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(
        json.dumps(items, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    morning_count = sum(
        1
        for item in items
        if "pagi" in item["sessions"] and item["morningOrder"] is not None
    )
    evening_count = sum(
        1
        for item in items
        if "petang" in item["sessions"] and item["eveningOrder"] is not None
    )
    hundred_count = sum(1 for item in items if item["count"] and item["count"] >= 100)

    print("Dzikir data generated")
    print(f"- rows: {len(items)}")
    print(f"- morning items: {morning_count}")
    print(f"- evening items: {evening_count}")
    print(f"- 100x counter items: {hundred_count}")
    if warnings:
        print("- warnings:")
        for warning in warnings:
            print(f"  - {warning}")
    else:
        print("- warnings: none")

    return items, warnings


def load_font(size):
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Bold.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def build_icon(size):
    bg = "#F8F3EA"
    surface = "#FFFDF7"
    ink = "#2F2A22"
    accent = "#4F5D3A"
    terracotta = "#B9865B"
    soft = "#E8DDCC"

    image = Image.new("RGB", (size, size), bg)
    draw = ImageDraw.Draw(image)

    margin = int(size * 0.12)
    shadow = int(size * 0.045)
    radius = int(size * 0.08)
    box = [margin, margin, size - margin - shadow, size - margin - shadow]
    shadow_box = [box[0] + shadow, box[1] + shadow, box[2] + shadow, box[3] + shadow]

    draw.rounded_rectangle(shadow_box, radius=radius, fill=ink)
    draw.rounded_rectangle(box, radius=radius, fill=surface, outline=ink, width=max(4, size // 48))

    band_height = int(size * 0.18)
    draw.rounded_rectangle(
        [box[0], box[1], box[2], box[1] + band_height],
        radius=radius,
        fill=accent,
        outline=ink,
        width=max(4, size // 48),
    )
    draw.rectangle(
        [box[0], box[1] + band_height - radius, box[2], box[1] + band_height],
        fill=accent,
    )
    draw.line(
        [box[0], box[1] + band_height, box[2], box[1] + band_height],
        fill=ink,
        width=max(4, size // 48),
    )

    font = load_font(int(size * 0.28))
    small_font = load_font(int(size * 0.075))
    text = "DP"
    text_box = draw.textbbox((0, 0), text, font=font)
    text_width = text_box[2] - text_box[0]
    text_height = text_box[3] - text_box[1]
    text_x = (size - text_width) / 2
    text_y = box[1] + band_height + int(size * 0.12)
    draw.text((text_x, text_y), text, font=font, fill=ink)

    pill = [int(size * 0.27), int(size * 0.68), int(size * 0.73), int(size * 0.79)]
    draw.rounded_rectangle(
        pill,
        radius=int(size * 0.035),
        fill=soft,
        outline=ink,
        width=max(3, size // 64),
    )

    label = "DZIKIR"
    label_box = draw.textbbox((0, 0), label, font=small_font)
    label_width = label_box[2] - label_box[0]
    label_height = label_box[3] - label_box[1]
    draw.text(
        ((size - label_width) / 2, pill[1] + ((pill[3] - pill[1] - label_height) / 2) - 1),
        label,
        font=small_font,
        fill=ink,
    )

    dot_radius = int(size * 0.025)
    draw.ellipse(
        [
            int(size * 0.74),
            int(size * 0.27),
            int(size * 0.74) + dot_radius * 2,
            int(size * 0.27) + dot_radius * 2,
        ],
        fill=terracotta,
        outline=ink,
        width=max(2, size // 96),
    )

    return image


def build_icons():
    icon_dir = ROOT / "icons"
    icon_dir.mkdir(parents=True, exist_ok=True)
    for size in (192, 512):
        build_icon(size).save(icon_dir / f"icon-{size}.png")
    print("PWA icons generated")
    print("- icons/icon-192.png")
    print("- icons/icon-512.png")


if __name__ == "__main__":
    build_dzikir_json()
    build_icons()
